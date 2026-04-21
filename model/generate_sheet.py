"""
Generate the MagCry solver spreadsheet.

Pure Excel formulas — no macros. Open it during the game,
fill in yellow cells, get instant EV/decision/inference output.

Usage:
    python generate_sheet.py
    -> produces trading_game_solver.xlsx in this directory
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

# --- Constants ---
DECK = [-10, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 20]
DECK_SUM = 130
N_CARDS = 17
N_OPPONENTS = 4

# --- Styles ---
YELLOW_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")
BLUE_FILL = PatternFill(start_color="FFCFE2F3", end_color="FFCFE2F3", fill_type="solid")
GREY_FILL = PatternFill(start_color="FFF3F3F3", end_color="FFF3F3F3", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FF4A86C8", end_color="FF4A86C8", fill_type="solid")

BOLD = Font(bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)
INPUT_FONT = Font(size=12)
OUTPUT_FONT = Font(bold=True, size=12)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_game_sheet(wb):
    """Sheet 1: GAME — main input/output interface."""
    ws = wb.active
    ws.title = "GAME"

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    # === TITLE ===
    ws["B1"] = "TRADING GAME SOLVER"
    ws["B1"].font = TITLE_FONT

    # === YOUR INPUTS ===
    ws["B3"] = "YOUR INPUTS"
    ws["B3"].font = SECTION_FONT

    labels = ["Your Card:", "Central 1:", "Central 2:", "Central 3:"]
    input_cells = ["C4", "C5", "C6", "C7"]
    for i, (label, cell) in enumerate(zip(labels, input_cells)):
        ws[f"B{4+i}"] = label
        ws[f"B{4+i}"].font = BOLD
        ws[cell].fill = YELLOW_FILL
        ws[cell].font = INPUT_FONT
        ws[cell].border = THIN_BORDER

    # === SETTINGS ===
    ws["E3"] = "SETTINGS"
    ws["E3"].font = SECTION_FONT
    ws["E4"] = "Bluff Prior (p):"
    ws["E4"].font = BOLD
    ws["F4"] = 0.7
    ws["F4"].fill = YELLOW_FILL
    ws["F4"].border = THIN_BORDER
    ws["F4"].number_format = "0.0"
    ws["E5"] = "Sigma:"
    ws["E5"].font = BOLD
    ws["F5"] = 3
    ws["F5"].fill = YELLOW_FILL
    ws["F5"].border = THIN_BORDER

    # === YOUR EV (auto-calculated) ===
    row = 9
    ws[f"B{row}"] = "YOUR EV"
    ws[f"B{row}"].font = SECTION_FONT

    # Helper calcs (hidden logic)
    # KnownCount = 1 + number of non-blank centrals
    # KnownSum = YourCard + sum of centrals
    # UnknownCount = 7 - COUNTA(centrals)
    # RemainingMean = (130 - KnownSum) / (17 - KnownCount)
    # FairTotal = KnownSum + UnknownCount * RemainingMean

    row = 10
    ws[f"B{row}"] = "Known Cards:"
    ws[f"C{row}"] = '=1+COUNTA(C5:C7)'
    ws[f"C{row}"].fill = GREY_FILL

    row = 11
    ws[f"B{row}"] = "Known Sum:"
    ws[f"C{row}"] = '=C4+SUM(C5:C7)'
    ws[f"C{row}"].fill = GREY_FILL

    row = 12
    ws[f"B{row}"] = "Unknowns in Play:"
    ws[f"C{row}"] = '=7-COUNTA(C5:C7)'
    ws[f"C{row}"].fill = GREY_FILL

    row = 13
    ws[f"B{row}"] = "Remaining Pool Mean:"
    ws[f"C{row}"] = '=(130-C11)/(17-C10)'
    ws[f"C{row}"].fill = GREY_FILL
    ws[f"C{row}"].number_format = "0.00"

    row = 15
    ws[f"B{row}"] = "Fair Total (EV):"
    ws[f"B{row}"].font = BOLD
    ws[f"C{row}"] = '=IF(C4="","",C11+C12*C13)'
    ws[f"C{row}"].fill = GREEN_FILL
    ws[f"C{row}"].font = OUTPUT_FONT
    ws[f"C{row}"].number_format = "0.0"

    row = 16
    ws[f"B{row}"] = "Fair Mid (rounded):"
    ws[f"C{row}"] = '=IF(C15="","",ROUND(C15,0))'
    ws[f"C{row}"].fill = GREEN_FILL
    ws[f"C{row}"].font = OUTPUT_FONT

    row = 17
    ws[f"B{row}"] = "Your Bid:"
    ws[f"C{row}"] = '=IF(C16="","",C16-1)'
    ws[f"C{row}"].fill = GREEN_FILL
    ws[f"C{row}"].font = OUTPUT_FONT

    row = 18
    ws[f"B{row}"] = "Your Ask:"
    ws[f"C{row}"] = '=IF(C16="","",C16+1)'
    ws[f"C{row}"].fill = GREEN_FILL
    ws[f"C{row}"].font = OUTPUT_FONT

    # === DECISION HELPER ===
    row = 20
    ws[f"B{row}"] = "DECISION HELPER"
    ws[f"B{row}"].font = SECTION_FONT

    row = 21
    ws[f"B{row}"] = "Offered Bid:"
    ws[f"B{row}"].font = BOLD
    ws[f"C{row}"] = None
    ws[f"C{row}"].fill = YELLOW_FILL
    ws[f"C{row}"].border = THIN_BORDER

    row = 22
    ws[f"B{row}"] = "Offered Ask:"
    ws[f"B{row}"].font = BOLD
    ws[f"C{row}"] = None
    ws[f"C{row}"].fill = YELLOW_FILL
    ws[f"C{row}"].border = THIN_BORDER

    row = 24
    ws[f"B{row}"] = "Buy EV:"
    ws[f"B{row}"].font = BOLD
    ws[f"C{row}"] = '=IF(OR(C15="",C22=""),"",C15-C22)'
    ws[f"C{row}"].fill = GREEN_FILL
    ws[f"C{row}"].font = OUTPUT_FONT
    ws[f"C{row}"].number_format = "+0.0;-0.0;0"

    row = 25
    ws[f"B{row}"] = "Sell EV:"
    ws[f"B{row}"].font = BOLD
    ws[f"C{row}"] = '=IF(OR(C15="",C21=""),"",C21-C15)'
    ws[f"C{row}"].fill = GREEN_FILL
    ws[f"C{row}"].font = OUTPUT_FONT
    ws[f"C{row}"].number_format = "+0.0;-0.0;0"

    row = 27
    ws[f"B{row}"] = "RECOMMENDATION:"
    ws[f"B{row}"].font = Font(bold=True, size=12)
    # Logic: if both blank -> "". If buy > 0 and buy >= sell -> BUY. If sell > 0 -> SELL. Else WALK.
    ws[f"C{row}"] = (
        '=IF(OR(C24="",C25=""),"—",'
        'IF(AND(C24>0,C24>=C25),"BUY",'
        'IF(C25>0,"SELL","WALK")))'
    )
    ws[f"C{row}"].fill = GREEN_FILL
    ws[f"C{row}"].font = Font(bold=True, size=14)

    # === OPPONENT SUMMARY ===
    row = 29
    ws[f"B{row}"] = "OPPONENT INFERENCE"
    ws[f"B{row}"].font = SECTION_FONT

    row = 30
    headers = ["Player", "Obs. Mid", "Likely Card", "Confidence"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=2 + i)
        cell.value = h
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for opp in range(N_OPPONENTS):
        r = 31 + opp
        ws.cell(row=r, column=2).value = f"Opponent {opp + 1}"
        ws.cell(row=r, column=2).font = BOLD

        # Observed mid - input cell
        mid_cell = ws.cell(row=r, column=3)
        mid_cell.fill = YELLOW_FILL
        mid_cell.border = THIN_BORDER

        # Likely Card - pulls from INFERENCE sheet
        # Each opponent block: header at row 2+opp*19, col headers at +1, data at +2
        # Data rows: (2 + opp*19 + 2) to (2 + opp*19 + 2 + 16) = 17 card rows
        inf_start = 4 + opp * 19
        inf_end = inf_start + 16  # 17 cards
        card_cell = ws.cell(row=r, column=4)
        card_cell.value = (
            f'=IF(C{r}="","—",'
            f'INDEX(INFERENCE!A{inf_start}:A{inf_end},'
            f'MATCH(MAX(INFERENCE!G{inf_start}:G{inf_end}),'
            f'INFERENCE!G{inf_start}:G{inf_end},0)))'
        )
        card_cell.fill = GREEN_FILL
        card_cell.font = OUTPUT_FONT

        # Confidence - max probability
        conf_cell = ws.cell(row=r, column=5)
        conf_cell.value = (
            f'=IF(C{r}="","—",'
            f'MAX(INFERENCE!G{inf_start}:G{inf_end}))'
        )
        conf_cell.fill = GREEN_FILL
        conf_cell.font = OUTPUT_FONT
        conf_cell.number_format = "0%"

    return ws


def create_inference_sheet(wb):
    """Sheet 2: INFERENCE — Bayesian card probability calculations."""
    ws = wb.create_sheet("INFERENCE")

    ws.column_dimensions["A"].width = 6   # Card
    ws.column_dimensions["B"].width = 10  # Possible?
    ws.column_dimensions["C"].width = 14  # Their EV (naive)
    ws.column_dimensions["D"].width = 14  # Their EV (bluff)
    ws.column_dimensions["E"].width = 14  # Naive Likelihood
    ws.column_dimensions["F"].width = 14  # Bluff Likelihood
    ws.column_dimensions["G"].width = 12  # P(card)

    # Reference to GAME sheet cells
    # Your card: GAME!C4
    # Centrals: GAME!C5, GAME!C6, GAME!C7
    # p (bluff prior): GAME!F4
    # sigma: GAME!F5
    # Opponent mids: GAME!C31, C32, C33, C34

    for opp in range(N_OPPONENTS):
        block_start = 2 + opp * 19  # row where this opponent's block starts
        header_row = block_start

        # Opponent header
        ws.cell(row=header_row, column=1).value = f"Opponent {opp + 1}"
        ws.cell(row=header_row, column=1).font = SECTION_FONT

        # Observed mid reference
        opp_mid_ref = f"GAME!C{31 + opp}"
        ws.cell(row=header_row, column=3).value = f"Obs Mid → {opp_mid_ref}"
        ws.cell(row=header_row, column=3).font = Font(italic=True, size=9)

        # Column headers
        col_headers = ["Card", "Poss?", "Naive EV", "Bluff EV", "L(naive)", "L(bluff)", "P(card)"]
        data_start = header_row + 1
        for i, h in enumerate(col_headers):
            cell = ws.cell(row=data_start, column=1 + i)
            cell.value = h
            cell.font = BOLD
            cell.fill = GREY_FILL

        # 17 rows, one per deck card
        for card_idx, card_val in enumerate(DECK):
            r = data_start + 1 + card_idx

            # Column A: Card value
            ws.cell(row=r, column=1).value = card_val

            # Column B: Possible? (not your card, not a revealed central)
            # =AND(A{r}<>GAME!C4, OR(GAME!C5="",A{r}<>GAME!C5), OR(GAME!C6="",A{r}<>GAME!C6), OR(GAME!C7="",A{r}<>GAME!C7))
            ws.cell(row=r, column=2).value = (
                f'=AND(A{r}<>GAME!C4,'
                f'OR(GAME!C5="",A{r}<>GAME!C5),'
                f'OR(GAME!C6="",A{r}<>GAME!C6),'
                f'OR(GAME!C7="",A{r}<>GAME!C7))'
            )

            # Column C: Their EV if they hold this card (naive - they quote true EV)
            # their_ev = card + SUM(centrals) + unknowns * (130 - card - SUM(centrals)) / (16 - n_revealed)
            # unknowns = 7 - COUNTA(GAME!C5:C7)
            # n_revealed = COUNTA(GAME!C5:C7)
            ws.cell(row=r, column=3).value = (
                f'=A{r}+SUM(GAME!C5:C7)'
                f'+(7-COUNTA(GAME!C5:C7))'
                f'*(130-A{r}-SUM(GAME!C5:C7))'
                f'/(16-COUNTA(GAME!C5:C7))'
            )
            ws.cell(row=r, column=3).number_format = "0.0"

            # Column D: Bluff EV (if card > mean ~7.65, they quote LOW; if below, quote HIGH)
            # bluff_offset from REFERENCE sheet (cell B5)
            # =IF(A{r}>7.65, C{r}-REFERENCE!B5, C{r}+REFERENCE!B5)
            ws.cell(row=r, column=4).value = (
                f'=IF(A{r}>7.65,C{r}-REFERENCE!B5,C{r}+REFERENCE!B5)'
            )
            ws.cell(row=r, column=4).number_format = "0.0"

            # Column E: Naive Likelihood = IF(possible AND obs_mid not blank, EXP(-(naive_ev - obs_mid)^2 / (2*sigma^2)), 0)
            ws.cell(row=r, column=5).value = (
                f'=IF(AND(B{r},{opp_mid_ref}<>""),'
                f'EXP(-((C{r}-{opp_mid_ref})^2)/(2*GAME!F5^2)),0)'
            )
            ws.cell(row=r, column=5).number_format = "0.0000"

            # Column F: Bluff Likelihood
            ws.cell(row=r, column=6).value = (
                f'=IF(AND(B{r},{opp_mid_ref}<>""),'
                f'EXP(-((D{r}-{opp_mid_ref})^2)/(2*GAME!F5^2)),0)'
            )
            ws.cell(row=r, column=6).number_format = "0.0000"

            # Column G: P(card) = combined likelihood normalized
            # Combined = p * naive + (1-p) * bluff
            # Then normalize by sum of all combined likelihoods for this opponent
            # First compute combined in-place in the formula
            like_range_e = f"E{data_start + 1}:E{data_start + 17}"
            like_range_f = f"F{data_start + 1}:F{data_start + 17}"
            ws.cell(row=r, column=7).value = (
                f'=IF({opp_mid_ref}="",'
                f'IF(B{r},1/COUNTIF(B{data_start+1}:B{data_start+17},TRUE),0),'
                f'IF(B{r},'
                f'(GAME!F4*E{r}+(1-GAME!F4)*F{r})'
                f'/SUMPRODUCT(GAME!F4*{like_range_e}+(1-GAME!F4)*{like_range_f})'
                f',0))'
            )
            ws.cell(row=r, column=7).number_format = "0.0%"
            ws.cell(row=r, column=7).fill = BLUE_FILL

    return ws


def create_reference_sheet(wb):
    """Sheet 3: REFERENCE — deck data and settings."""
    ws = wb.create_sheet("REFERENCE")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12

    ws["A1"] = "REFERENCE DATA"
    ws["A1"].font = TITLE_FONT

    # Settings
    ws["A3"] = "Settings"
    ws["A3"].font = SECTION_FONT
    ws["A4"] = "Sigma:"
    ws["B4"] = "=GAME!F5"
    ws["A5"] = "Bluff Offset:"
    ws["B5"] = 6
    ws["B5"].fill = YELLOW_FILL
    ws["B5"].border = THIN_BORDER
    ws["A6"] = "Bluff Prior (p):"
    ws["B6"] = "=GAME!F4"

    # Deck
    ws["A8"] = "Deck"
    ws["A8"].font = SECTION_FONT
    ws["A9"] = "Card"
    ws["B9"] = "Mean Excl."
    ws["A9"].font = BOLD
    ws["B9"].font = BOLD

    for i, card in enumerate(DECK):
        ws.cell(row=10 + i, column=1).value = card
        # Mean of remaining 16 cards if this card is removed
        ws.cell(row=10 + i, column=2).value = round((DECK_SUM - card) / 16, 2)

    # Summary stats
    ws["D3"] = "Deck Stats"
    ws["D3"].font = SECTION_FONT
    ws["D4"] = "Total Cards:"
    ws["E4"] = N_CARDS
    ws["D5"] = "Deck Sum:"
    ws["E5"] = DECK_SUM
    ws["D6"] = "Mean per Card:"
    ws["E6"] = round(DECK_SUM / N_CARDS, 2)
    ws["D7"] = "Cards in Play:"
    ws["E7"] = 8
    ws["D8"] = "Players:"
    ws["E8"] = 5

    # How to use
    ws["A29"] = "HOW TO USE"
    ws["A29"].font = SECTION_FONT
    ws["A30"] = "1. Enter your card in the yellow cell on GAME sheet"
    ws["A31"] = "2. As central cards are revealed, enter them"
    ws["A32"] = "3. When an opponent quotes, enter their mid-price"
    ws["A33"] = "4. When offered a price, enter bid/ask in Decision Helper"
    ws["A34"] = "5. The sheet tells you: fair price, who has what, buy/sell/walk"
    ws["A36"] = "p=1.0 → assume no bluffing. p=0.5 → uncertain. p=0.0 → everyone bluffs."
    ws["A37"] = "Sigma: lower = more confident inference (try 2-4)"
    ws["A38"] = "Bluff Offset: how far bluffers shift from true EV (try 4-8)"

    return ws


def main():
    wb = Workbook()
    create_game_sheet(wb)
    create_inference_sheet(wb)
    create_reference_sheet(wb)

    output_path = "/Users/cgilmart/github/conor-random-stuff/Ai/test-challenge/gary-game/model/trading_game_solver.xlsx"
    wb.save(output_path)
    print(f"Generated: {output_path}")
    print("Open in Excel/Numbers/Google Sheets and fill in the yellow cells.")


if __name__ == "__main__":
    main()
